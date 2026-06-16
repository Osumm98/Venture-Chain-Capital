"""Extract all member data from Excel into a structured JSON."""
import openpyxl
import json

wb = openpyxl.load_workbook(
    r"c:\Users\mrmaf\OneDrive\Desktop\VCC Landing\Data\VCC 2026_ACCOUNTS, TOKENS AND MEMBERS.xlsx",
    data_only=True,
)

# ── Get all member sheet names ──
member_sheets = [s for s in wb.sheetnames if s.startswith("BWG") and "TEMPLATE" not in s]

# ── Extract Instalment Tokens ──
inst_ws = wb["Instalment Tokens"]
tiers_inst = ["ENTRY", "SILVER", "GOLD", "PLATINUM", "DIAMOND", "GROUP_1", "GROUP_2", "GROUP_3"]
# Header row is 5, data starts row 6
instalment_data = {}
for row in inst_ws.iter_rows(min_row=6, max_row=50, values_only=False):
    membership_no = row[3].value  # Column D
    if not membership_no:
        continue
    membership_no = str(membership_no).strip()
    tokens = {}
    for i, tier in enumerate(tiers_inst):
        val = row[4 + i].value  # Columns E through L
        tokens[tier] = int(val) if val else 0
    instalment_data[membership_no] = tokens

# ── Extract Carry-Over Tokens ──
carry_ws = wb["Carry-Over Tokens"]
carry_data = {}
for row in carry_ws.iter_rows(min_row=6, max_row=50, values_only=False):
    membership_no = row[3].value
    if not membership_no:
        continue
    membership_no = str(membership_no).strip()
    tokens = {}
    for i, tier in enumerate(tiers_inst):
        val = row[4 + i].value
        tokens[tier] = int(val) if val else 0
    carry_data[membership_no] = tokens

# ── Extract Total Active Tokens ──
total_ws = wb["Total Active Tokens"]
total_data = {}
for row in total_ws.iter_rows(min_row=6, max_row=50, values_only=False):
    membership_no = row[3].value
    if not membership_no:
        continue
    membership_no = str(membership_no).strip()
    tokens = {}
    for i, tier in enumerate(tiers_inst):
        val = row[4 + i].value
        tokens[tier] = int(val) if val else 0
    total_data[membership_no] = tokens

# ── Extract Token Cashout Prices (current) ──
cash_ws = wb["Weekly Token Cashout Prices"]
cashout_prices = {}
# Row 7 has the current prices
for i, tier in enumerate(tiers_inst):
    val = cash_ws.cell(row=7, column=2 + i).value
    cashout_prices[tier] = round(float(val), 2) if val else 0.0

# ── Extract per-member sheet data ──
members = []
for sheet_name in member_sheets:
    ws = wb[sheet_name]
    member = {}
    
    # Extract membership number from cell C6
    member["membershipNo"] = str(ws.cell(row=6, column=3).value or "").strip()
    
    # Extract name from cell C5
    full_name = str(ws.cell(row=5, column=3).value or "").strip()
    parts = full_name.rsplit(" ", 1) if full_name else ["", ""]
    if len(parts) == 2:
        member["firstName"] = parts[0]
        member["lastName"] = parts[1]
    else:
        member["firstName"] = full_name
        member["lastName"] = ""
    
    # Date of membership
    member["dateOfMembership"] = str(ws.cell(row=7, column=3).value or "")
    
    # Monthly contribution
    member["monthlyContribution"] = ws.cell(row=22, column=3).value or 0
    
    # Account status and value
    member["accountStatus"] = str(ws.cell(row=24, column=3).value or "ACTIVE")
    member["accountValue"] = ws.cell(row=24, column=4).value or 0
    
    # Token holdings from member sheet
    member["tokens"] = {}
    tier_names = ["ENTRY", "SILVER", "GOLD", "PLATINUM", "DIAMOND", "GROUP_1", "GROUP_2", "GROUP_3"]
    for idx, tier in enumerate(tier_names):
        row_num = 13 + idx  # ENTRY at row 13, SILVER at 14, etc.
        instalment_count = ws.cell(row=row_num, column=3).value or 0
        carry_count = ws.cell(row=row_num, column=4).value or 0
        total_count = ws.cell(row=row_num, column=5).value or 0
        cashout_price = ws.cell(row=row_num, column=7).value or 0
        
        member["tokens"][tier] = {
            "instalment": int(instalment_count) if instalment_count else 0,
            "carryOver": int(carry_count) if carry_count else 0,
            "total": int(total_count) if total_count else 0,
            "cashoutPrice": round(float(cashout_price), 2) if cashout_price else 0.0,
        }
    
    # Account contributions and fees
    member["accountContributions"] = ws.cell(row=27, column=3).value or 0
    member["accountFees"] = ws.cell(row=28, column=3).value or 0
    member["accountContributionsAfterFees"] = ws.cell(row=29, column=3).value or 0
    
    # Instalment tokens detail
    member["instalmentTokenDetails"] = []
    for row in ws.iter_rows(min_row=33, max_row=55, values_only=False):
        token_held = row[1].value  # Column B
        token_number = row[2].value  # Column C
        instalments_paid = row[3].value  # Column D
        if token_number and str(token_number).startswith("VCC"):
            tier_label = str(row[0].value or "").strip()
            member["instalmentTokenDetails"].append({
                "tier": tier_label if tier_label else None,
                "tokenNumber": str(token_number),
                "instalmentsPaid": int(instalments_paid) if instalments_paid else 0,
            })
    
    # Extract Payments from Account Notes
    member["payments"] = []
    # Payments typically start around row 75
    for row in ws.iter_rows(min_row=75, max_row=100, values_only=True):
        if row[0] and str(row[0]).startswith("PAYMENT"):
            if row[2] is not None:
                member["payments"].append({
                    "reference": str(row[0]),
                    "date": str(row[1]) if row[1] else None,
                    "amount": round(float(row[2]), 2) if row[2] else 0.0
                })
    
    members.append(member)

# ── Token Installments and Fees ──
fees_ws = wb["Token Instalments and Fees"]
subscription_prices = {}
for i, tier in enumerate(tiers_inst):
    val = fees_ws.cell(row=6, column=2 + i).value
    subscription_prices[tier] = round(float(val), 2) if val else 0.0

admin_fees = {}
for i, tier in enumerate(tiers_inst):
    val = fees_ws.cell(row=7, column=2 + i).value
    admin_fees[tier] = round(float(val) * 100, 2) if val else 0.0

profit_fees = {}
for i, tier in enumerate(tiers_inst):
    val = fees_ws.cell(row=8, column=2 + i).value
    profit_fees[tier] = round(float(val) * 100, 2) if val else 0.0

output = {
    "subscriptionPrices": subscription_prices,
    "adminFees": admin_fees,
    "profitFees": profit_fees,
    "cashoutPrices": cashout_prices,
    "instalmentTokens": instalment_data,
    "carryOverTokens": carry_data,
    "totalActiveTokens": total_data,
    "members": members,
}

print(json.dumps(output, indent=2, default=str))
